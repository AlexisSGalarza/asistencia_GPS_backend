from rest_framework.views import APIView
from django.http import FileResponse, HttpResponse
from django.utils import timezone as tz
from common.permissions import EsSupervisorOAdmin
from common.reportes import generar_reporte_asistencia, generar_reporte_incidencias
from apps.locations.models import Asistencia, Incidencia
import io


class ReporteAsistenciaPDFView(APIView):
    """
    GET /api/reportes/asistencia/?fecha_inicio=2026-01-01&fecha_fin=2026-01-31&usuario=1
    Descarga un PDF con el reporte de asistencias.
    """
    permission_classes = [EsSupervisorOAdmin]

    def get(self, request):
        qs = Asistencia.objects.select_related('usuario', 'perimetro').all()

        fecha_inicio = request.query_params.get('fecha_inicio')
        fecha_fin = request.query_params.get('fecha_fin')
        usuario_id = request.query_params.get('usuario')

        if usuario_id:
            qs = qs.filter(usuario_id=usuario_id)
        if fecha_inicio:
            qs = qs.filter(fecha_hora__date__gte=fecha_inicio)
        if fecha_fin:
            qs = qs.filter(fecha_hora__date__lte=fecha_fin)

        qs = qs.order_by('fecha_hora')

        buffer = generar_reporte_asistencia(
            list(qs),
            fecha_inicio=fecha_inicio,
            fecha_fin=fecha_fin,
        )

        return FileResponse(
            buffer,
            as_attachment=True,
            filename=f'reporte_asistencia_{fecha_inicio or "completo"}.pdf',
            content_type='application/pdf',
        )


class ReporteIncidenciasPDFView(APIView):
    """
    GET /api/reportes/incidencias/?fecha_inicio=2026-01-01&fecha_fin=2026-01-31&usuario=1
    Descarga un PDF con el reporte de incidencias.
    """
    permission_classes = [EsSupervisorOAdmin]

    def get(self, request):
        qs = Incidencia.objects.select_related('usuario').all()

        fecha_inicio = request.query_params.get('fecha_inicio')
        fecha_fin = request.query_params.get('fecha_fin')
        usuario_id = request.query_params.get('usuario')

        if usuario_id:
            qs = qs.filter(usuario_id=usuario_id)
        if fecha_inicio:
            qs = qs.filter(fecha__gte=fecha_inicio)
        if fecha_fin:
            qs = qs.filter(fecha__lte=fecha_fin)

        # 'olvido_salida' es informativo (no afecta nómina); se excluye del reporte
        # El admin puede consultarlas aparte con ?tipo=olvido_salida en el API.
        qs = qs.exclude(tipo=Incidencia.Tipo.OLVIDO_SALIDA)
        qs = qs.order_by('fecha')

        buffer = generar_reporte_incidencias(
            list(qs),
            fecha_inicio=fecha_inicio,
            fecha_fin=fecha_fin,
        )

        return FileResponse(
            buffer,
            as_attachment=True,
            filename=f'reporte_incidencias_{fecha_inicio or "completo"}.pdf',
            content_type='application/pdf',
        )


class ReporteAsistenciaExcelView(APIView):
    """
    GET /api/reportes/asistencia/excel/?fecha_inicio=2026-01-01&fecha_fin=2026-01-31&usuario=1
    Descarga un Excel con el reporte de asistencias.
    """
    permission_classes = [EsSupervisorOAdmin]

    def get(self, request):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            return HttpResponse('openpyxl no instalado. Ejecuta: pip install openpyxl', status=500)

        qs = Asistencia.objects.select_related('usuario', 'perimetro').all()

        fecha_inicio = request.query_params.get('fecha_inicio')
        fecha_fin = request.query_params.get('fecha_fin')
        usuario_id = request.query_params.get('usuario')

        if usuario_id:
            qs = qs.filter(usuario_id=usuario_id)
        if fecha_inicio:
            qs = qs.filter(fecha_hora__date__gte=fecha_inicio)
        if fecha_fin:
            qs = qs.filter(fecha_hora__date__lte=fecha_fin)

        qs = qs.order_by('fecha_hora')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Asistencia'

        header_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)

        headers = ['#', 'Maestro', 'Correo', 'Tipo', 'Fecha/Hora', 'Válido', 'Distancia (m)', 'Perímetro']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        for i, a in enumerate(qs, 1):
            # Convertir a hora local México antes de formatear
            fecha_local = tz.localtime(a.fecha_hora)
            ws.append([
                i,
                a.usuario.nombre,
                a.usuario.correo,
                a.get_tipo_display(),
                fecha_local.strftime('%d/%m/%Y %H:%M'),
                'Sí' if a.valido else 'No',
                round(a.distancia_metros, 1),
                a.perimetro.nombre,
            ])

        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        filename = f'asistencia_{fecha_inicio or "completo"}.xlsx'
        response = HttpResponse(
            buffer.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


class ReporteIncidenciasExcelView(APIView):
    """
    GET /api/reportes/incidencias/excel/?fecha_inicio=2026-01-01&fecha_fin=2026-01-31&usuario=1
    Descarga un Excel con el reporte de incidencias.
    """
    permission_classes = [EsSupervisorOAdmin]

    def get(self, request):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            return HttpResponse('openpyxl no instalado. Ejecuta: pip install openpyxl', status=500)

        qs = Incidencia.objects.select_related('usuario').all()

        fecha_inicio = request.query_params.get('fecha_inicio')
        fecha_fin = request.query_params.get('fecha_fin')
        usuario_id = request.query_params.get('usuario')

        if usuario_id:
            qs = qs.filter(usuario_id=usuario_id)
        if fecha_inicio:
            qs = qs.filter(fecha__gte=fecha_inicio)
        if fecha_fin:
            qs = qs.filter(fecha__lte=fecha_fin)

        # 'olvido_salida' es informativo (no afecta nómina); se excluye del reporte
        # El admin puede consultarlas aparte con ?tipo=olvido_salida en el API.
        qs = qs.exclude(tipo=Incidencia.Tipo.OLVIDO_SALIDA)
        qs = qs.order_by('fecha')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Incidencias'

        header_fill = PatternFill(start_color='C0392B', end_color='C0392B', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)

        headers = ['#', 'Maestro', 'Correo', 'Tipo', 'Fecha', 'Descripción']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        for i, inc in enumerate(qs, 1):
            ws.append([
                i,
                inc.usuario.nombre,
                inc.usuario.correo,
                inc.get_tipo_display(),
                inc.fecha.strftime('%d/%m/%Y'),
                inc.descripcion,
            ])

        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        filename = f'incidencias_{fecha_inicio or "completo"}.xlsx'
        response = HttpResponse(
            buffer.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
